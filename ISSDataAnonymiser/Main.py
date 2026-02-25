from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def get_valid_column(ws, prompt_text):
    while True:
        col = input(prompt_text).strip().upper()

        try:
            col_number = column_index_from_string(col)
        except ValueError:
            print("Invalid column format. Please enter letters only (example: A, B, AA).")
            continue

        if col_number <= ws.max_column:
            return col
        else:
            print("Column does not exist in this sheet. Try again.")


# ----------- USER INPUT SECTION -----------
while True:
    filename = input("Enter Excel file name (example: data.xlsx): ")
    try:
        wb = load_workbook(filename)
        break
    except:
        print("File not found. Perhaps check spelling or check that it is in correct folder")
ws = wb.active

first_col = get_valid_column(ws, "Enter FIRST name column letter:")
last_col = get_valid_column(ws, "Enter LAST name column letter:")
email_col = get_valid_column(ws, "Enter EMAIL column letter:")



# ----------- CONFIRM STAGE -----------
print(f"You selected:\n File:{filename}\n First Name Column{first_col}\n Last Name Column{last_col}\n Email Column{email_col}")

while True:
    confirm = input("Proceed?(y/n)").strip().upper()
    if confirm == ("Y"):
        break
    else:
        print("Please restart program")
        quit()



# ----------- NAME ANONYMIZATION -----------
name_map = {}
email_map = {}
counter = 1
ecounter = 1

for row in range(2, ws.max_row + 1):  # skip header row
    first = ws[f"{first_col}{row}"].value
    last = ws[f"{last_col}{row}"].value
    email = ws[f"{email_col}{row}"].value

    if first and last:
        full_name = f"{str(first).strip()} {str(last).strip()}"

        if full_name not in name_map:
            name_map[full_name] = f"Person_{counter:03}"
            counter += 1

        anon_value = name_map[full_name]

        # Replace BOTH columns with same anonymized value
        ws[f"{first_col}{row}"] = anon_value
        ws[f"{last_col}{row}"] = anon_value

    if email:
        email_clean = str(email).strip()

        if email_clean not in email_map:
            email_map[email_clean] = f"email_{ecounter:03}"
            ecounter += 1

        eanon_value = email_map[email_clean]

        ws[f"{email_col}{row}"] = eanon_value


# ----------- SAVE OUTPUT -----------
output_name = "anonymized_" + filename
wb.save(output_name)

print(f"\nAnonymization complete! Saved as {output_name}")